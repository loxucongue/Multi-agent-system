"""Seed local development data for routes from an Excel source."""

from __future__ import annotations

import asyncio
import os
import re
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

from app.config.database import async_session_factory, engine
from app.models.database import Route, RoutePricing, RouteSchedule

_DEFAULT_EXCEL_PATH = Path(r"C:\Users\24159\Desktop\柯总-旅游智能体\8+51.xlsx")
_EXPECTED_COLUMNS = {
    "basic_info",
    "highlights",
    "itinerary_days",
    "cost_included",
    "notices",
    "file_url",
    "rag_abstract",
    "cost_excluded",
}


def _find_excel_path() -> Path:
    """Locate source Excel path from env, default path, or Desktop glob."""

    env_path = os.getenv("SEED_EXCEL_PATH", "").strip()
    if env_path:
        candidate = Path(env_path)
        if candidate.exists():
            return candidate

    if _DEFAULT_EXCEL_PATH.exists():
        return _DEFAULT_EXCEL_PATH

    desktop = Path.home() / "Desktop"
    matches = list(desktop.glob("*/8+51.xlsx"))
    if matches:
        return matches[0]

    raise FileNotFoundError(
        "Excel source not found. Set SEED_EXCEL_PATH or place 8+51.xlsx under Desktop subfolder."
    )


def _parse_title(basic_info: str) -> str:
    match = re.search(r"行程标题为[“\"](.*?)[”\"]", basic_info)
    if match:
        return match.group(1).strip()
    return basic_info[:80].strip() or "未命名线路"


def _parse_destinations(basic_info: str) -> list[str]:
    match = re.search(r"目的地为[“\"](.*?)[”\"]", basic_info)
    if not match:
        return []

    raw = match.group(1)
    parts = re.split(r"[、,，/]", raw)
    tags: list[str] = []
    for part in parts:
        tag = part.strip()
        if not tag:
            continue
        tag = re.sub(r"^中国-", "", tag)
        if tag not in tags:
            tags.append(tag)
    return tags


def _parse_days(basic_info: str) -> int:
    match = re.search(r"共(\d+)天\d+晚", basic_info)
    if not match:
        return 5
    return int(match.group(1))


def _parse_itinerary_json(itinerary_days: str) -> dict[str, Any]:
    """Build a compact structured itinerary while preserving original text."""

    day_items: list[dict[str, Any]] = []
    for day, title in re.findall(r"【第(\d+)天】[\s\S]*?- 行程标题/主题：([^\n]+)", itinerary_days):
        day_items.append({"day": int(day), "title": title.strip()})

    return {
        "days": day_items,
        "raw_text": itinerary_days,
    }


def _build_price(days: int) -> tuple[Decimal, Decimal]:
    price_min = Decimal("1299.00") + Decimal(days) * Decimal("700.00")
    price_max = price_min + Decimal("2200.00")
    return price_min, price_max


def _build_schedule(idx: int) -> list[dict[str, Any]]:
    today = date.today()
    offset = 10 + (idx % 20)
    return [
        {"date": (today + timedelta(days=offset)).isoformat(), "seats_left": max(6, 20 - (idx % 9))},
        {"date": (today + timedelta(days=offset + 14)).isoformat(), "seats_left": max(4, 16 - (idx % 7))},
    ]


def load_routes_from_excel() -> list[dict[str, Any]]:
    """Load and map Excel rows into Route seed payloads."""

    excel_path = _find_excel_path()
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    if not _EXPECTED_COLUMNS.issubset(set(headers)):
        missing = sorted(_EXPECTED_COLUMNS.difference(set(headers)))
        raise ValueError(f"Excel headers missing required columns: {missing}")

    index_map = {name: headers.index(name) for name in _EXPECTED_COLUMNS}

    routes: list[dict[str, Any]] = []
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        if row is None:
            continue

        def pick(column: str) -> str:
            value = row[index_map[column]]
            return str(value).strip() if value is not None else ""

        basic_info = pick("basic_info")
        if not basic_info:
            continue

        highlights = pick("highlights")
        itinerary_days = pick("itinerary_days")
        cost_included = pick("cost_included")
        notices = pick("notices")
        file_url = pick("file_url")
        rag_abstract = pick("rag_abstract")
        cost_excluded = pick("cost_excluded")

        title = _parse_title(basic_info)
        tags = _parse_destinations(basic_info)
        if not tags:
            tags = ["跟团游"]

        days = _parse_days(basic_info)
        price_min, price_max = _build_price(days)

        included_text = cost_included or "以合同与行前说明为准。"
        if cost_excluded:
            included_text = f"{included_text}\n\n费用不含：{cost_excluded}"

        routes.append(
            {
                "name": title,
                "supplier": "本地旅游供应",
                "tags": tags,
                "summary": rag_abstract or basic_info,
                "highlights": (
                    highlights
                    if highlights and highlights != "本行程未明确列出亮点"
                    else "行程亮点以最终出团通知与导游安排为准。"
                ),
                "base_info": basic_info,
                "itinerary_json": _parse_itinerary_json(itinerary_days),
                "notice": notices or "请以最终行前通知及签约条款为准。",
                "included": included_text,
                "doc_url": file_url or "https://example.com/routes/unknown.pdf",
                "is_hot": idx <= 8,
                "sort_weight": max(1, 120 - idx),
                "pricing": {"price_min": price_min, "price_max": price_max},
                "schedule": {"schedules_json": _build_schedule(idx)},
            }
        )

    return routes


async def seed_routes() -> None:
    """Insert seed routes with pricing and schedule records."""

    seed_routes_data = load_routes_from_excel()
    route_names = [item["name"] for item in seed_routes_data]
    inserted_count = 0

    try:
        async with async_session_factory() as session:
            existing_result = await session.execute(select(Route.name).where(Route.name.in_(route_names)))
            existing_names = set(existing_result.scalars().all())

            for item in seed_routes_data:
                if item["name"] in existing_names:
                    continue

                route = Route(
                    name=item["name"],
                    supplier=item["supplier"],
                    tags=item["tags"],
                    summary=item["summary"],
                    highlights=item["highlights"],
                    base_info=item["base_info"],
                    itinerary_json=item["itinerary_json"],
                    notice=item["notice"],
                    included=item["included"],
                    doc_url=item["doc_url"],
                    is_hot=item["is_hot"],
                    sort_weight=item["sort_weight"],
                )
                session.add(route)
                await session.flush()

                pricing = item["pricing"]
                schedule = item["schedule"]

                session.add(
                    RoutePricing(
                        route_id=route.id,
                        price_min=pricing["price_min"],
                        price_max=pricing["price_max"],
                        currency="CNY",
                    )
                )
                session.add(
                    RouteSchedule(
                        route_id=route.id,
                        schedules_json=schedule["schedules_json"],
                    )
                )
                inserted_count += 1

            await session.commit()
    finally:
        await engine.dispose()

    print(f"Seed completed. Loaded {len(seed_routes_data)} routes from Excel, inserted {inserted_count} new routes.")


if __name__ == "__main__":
    asyncio.run(seed_routes())
